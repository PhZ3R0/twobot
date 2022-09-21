from socket import timeout
import discord
from openpyxl import Workbook,load_workbook
from unicodedata import name
from discord.ext import commands

wb = load_workbook("Pasta.xlsx")
ws = wb.active

intents=discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix="!", description = "Nothing", intents=intents)

@bot.event
async def on_ready():
    print(f"Entrei como {bot.user}.")

segunda = []
@bot.command()
async def horario(ctx):
    await ctx.send("Como posso te ajudar?")
    diamsg = await bot.wait_for("message", timeout=30.0)
    segunda.append(diamsg.content)
    if diamsg.content.lower() == "seg" or diamsg.content.lower() == "segunda" or diamsg.content.lower() == "segunda-feira":
        await ctx.reply(f'''
                       Esses são os afazeres de SEGUNDA:
              {(ws["A4"].value)}
              {(ws["A5"].value)}
              {(ws["A6"].value)}
              {(ws["A7"].value)}
              {(ws["A8"].value)}
              {(ws["A9"].value)}
              {(ws["A10"].value)}
              {(ws["A11"].value)}
              {(ws["A12"].value)}
              {(ws["A13"].value)}
              {(ws["A14"].value)}
              {(ws["A15"].value)}
              {(ws["A16"].value)}
              {(ws["A17"].value)}
              {(ws["A18"].value)}
              ''')
    elif diamsg.content.lower() == "ter" or diamsg.content.lower() == "terça" or diamsg.content.lower() == "terça-feira":
            await ctx.reply(f'''
                       Esses são os afazeres de TERÇA-FEIRA:
              {(ws["C4"].value)}
              {(ws["C5"].value)}
              {(ws["C6"].value)}
              {(ws["C7"].value)}
              {(ws["C8"].value)}
              {(ws["C9"].value)}
              {(ws["C10"].value)}
              {(ws["C11"].value)}
              {(ws["C12"].value)}
              {(ws["C13"].value)}
              {(ws["C14"].value)}
              {(ws["C15"].value)}
              {(ws["C16"].value)}
              {(ws["C17"].value)}
              {(ws["C18"].value)}
              ''')
    elif diamsg.content.lower() == "qua" or diamsg.content.lower() == "quarta" or diamsg.content.lower() == "quarta-feira":
            await ctx.reply(f'''
                       Esses são os afazeres de QUARTA-FEIRA:
              {(ws["E4"].value)}
              {(ws["E5"].value)}
              {(ws["E6"].value)}
              {(ws["E7"].value)}
              {(ws["E8"].value)}
              {(ws["E9"].value)}
              {(ws["E10"].value)}
              {(ws["E11"].value)}
              {(ws["E12"].value)}
              {(ws["E13"].value)}
              {(ws["E14"].value)}
              {(ws["E15"].value)}
              {(ws["E16"].value)}
              {(ws["E17"].value)}
              {(ws["E18"].value)}
              ''')
    elif diamsg.content.lower() == "qui" or diamsg.content.lower() == "quinta" or diamsg.content.lower() == "quinta-feira":
            await ctx.reply(f'''
                       Esses são os afazeres de QUINTA-FEIRA:
              {(ws["G4"].value)}
              {(ws["G5"].value)}
              {(ws["G6"].value)}
              {(ws["G7"].value)}
              {(ws["G8"].value)}
              {(ws["G9"].value)}
              {(ws["G10"].value)}
              {(ws["G11"].value)}
              {(ws["G12"].value)}
              {(ws["G13"].value)}
              {(ws["G14"].value)}
              {(ws["G15"].value)}
              {(ws["G16"].value)}
              {(ws["G17"].value)}
              {(ws["G18"].value)}
              ''')
    elif diamsg.content.lower() == "sex" or diamsg.content.lower() == "sexta" or diamsg.content.lower() == "sexta-feira":
            await ctx.reply(f'''
                       Esses são os afazeres de SEXTA-FEIRA:
              {(ws["I4"].value)}
              {(ws["I5"].value)}
              {(ws["I6"].value)}
              {(ws["I7"].value)}
              {(ws["I8"].value)}
              {(ws["I9"].value)}
              {(ws["I10"].value)}
              {(ws["I11"].value)}
              {(ws["I12"].value)}
              {(ws["I13"].value)}
              {(ws["I14"].value)}
              {(ws["I15"].value)}
              {(ws["I16"].value)}
              {(ws["I17"].value)}
              {(ws["I18"].value)}
              ''')
    elif diamsg.content.lower() == "sab" or diamsg.content.lower() == "sabado":
            await ctx.reply(f'''
                       Esses são os afazeres de SABADO:
              {(ws["K4"].value)}
              {(ws["K5"].value)}
              {(ws["K6"].value)}
              {(ws["K7"].value)}
              {(ws["K8"].value)}
              {(ws["K9"].value)}
              {(ws["K10"].value)}
              {(ws["K11"].value)}
              {(ws["K12"].value)}
              {(ws["K13"].value)}
              {(ws["K14"].value)}
              {(ws["K15"].value)}
              {(ws["K16"].value)}
              {(ws["K17"].value)}
              {(ws["K18"].value)}
              ''')
    elif diamsg.content.lower() == "dom" or diamsg.content.lower() == "domingo":
            await ctx.reply(f'''
                       Esses são os afazeres de DOMINGO:
              {(ws["M4"].value)}
              {(ws["M5"].value)}
              {(ws["M6"].value)}
              {(ws["M7"].value)}
              {(ws["M8"].value)}
              {(ws["M9"].value)}
              {(ws["M10"].value)}
              {(ws["M11"].value)}
              {(ws["M12"].value)}
              {(ws["M13"].value)}
              {(ws["M14"].value)}
              {(ws["M15"].value)}
              {(ws["M16"].value)}
              {(ws["M17"].value)}
              {(ws["M18"].value)}
              ''')

@bot.command()
async def comandos(ctx):
    await ctx.reply('''
                    Aqui está minha lista de comandos:
                    !horario(respondo com seu horario em certo dia da semana)
                    !comandos(ajuda com os comandos)
                    !anotações(te respondo com suas anotações feitas no final da planilha!)
                    ''')
    
@bot.command()
async def anotações(ctx):
    await ctx.send(f'''
                       Essas são suas anotações:
              {(ws["O4"].value)}
              {(ws["O5"].value)}
              {(ws["O6"].value)}
              {(ws["O7"].value)}
              {(ws["O8"].value)}
              {(ws["O9"].value)}
              {(ws["O10"].value)}
              {(ws["O11"].value)}
              {(ws["O12"].value)}
              {(ws["O13"].value)}
              {(ws["O14"].value)}
              {(ws["O15"].value)}
              {(ws["O16"].value)}
              {(ws["O17"].value)}
              {(ws["O18"].value)}
              ''')







bot.run("")
